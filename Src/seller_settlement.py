# 2026-07-06 hoyeon.han 셀러 정산서 생성 비즈니스 로직 신규 작성
"""셀러 정산서(셀러별 정산내역서) 생성 모듈.

서버에 저장된 발주내역(.xlsm) 하나를 원본으로,
일자(기간/개별 날짜)·거래처·셀러(특이사항) 기준으로 필터한 데이터를
'품목별판매현황'/'품목별구매현황' 형식의 셀러 시트와 '상세내역' 시트로 구성된
.xlsx 정산서로 생성한다.
(2026-07-23 hoyeon.han: 매출/매입 모두 지원 — column_map=COLUMN_MAP_SALE/PURCHASE 로 전환)

발주내역 양식이 바뀌면 COLUMN_MAP_SALE 만 수정하면 되도록
모든 함수는 논리명(column_map 키) 기반으로 동작한다.

본 모듈의 모든 함수는 순수 함수(Streamlit/세션상태 의존 없음)로 작성한다.
"""

from __future__ import annotations

import os
from datetime import date
from typing import Any, Iterable

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 기존 모듈 재사용 (pages에서 sys.path에 Src가 추가되는 flat import 구조 —
# order_compare.py의 `from processing import to_num` 선례와 동일)
from order_compare import read_order_sheet
from processing import to_num
from settlement import (
    OUTPUT_COLS,
    SHEET_MAX_LEN,  # 2026-07-24 hoyeon.han: per_seller 상세시트명 31자 절단용
    TYPE_SALE,
    TYPE_PURCHASE,  # 2026-07-23 hoyeon.han: 매입 정산서 파일명(doc_type)용
    _sanitize_filename_part,
    _to_argb,
    apply_mapping,
    sanitize_sheet_name,
)
from settlement import build_period_line as _build_period_line_range

# ===== 상수 =====

# 논리명 → 발주내역 원본 컬럼명. 원본 양식이 바뀌면 이 dict만 수정한다.
COLUMN_MAP_SALE: dict[str, str] = {
    "date": "출고일",
    "vendor": "업체명",
    "category": "구분",
    "consignee": "수하인명",
    "address": "수하인주소",
    "product": "제품",
    "phone1": "연락처1",
    "phone2": "연락처2",
    "seller": "특이사항",  # 셀러(비고) 값
    "qty": "수량",
    "amount": "상품매출",
    "delivery": "판매배송비",
    "ferry": "도선료",
}

# 2026-07-23 hoyeon.han: 매입 정산서 컬럼 매핑 (기존 자리표시자 주석 → 실제 구현).
# 매출과 대칭이며 4개 키만 다르다(vendor/amount/delivery/ferry). ferry 의
# '매입도선료'는 원본 '도선료.1'을 load_orders 에서 정규화한 이름. 나머지는 동일.
COLUMN_MAP_PURCHASE: dict[str, str] = {
    "date": "출고일",
    "vendor": "매입처",
    "category": "구분",
    "consignee": "수하인명",
    "address": "수하인주소",
    "product": "제품",
    "phone1": "연락처1",
    "phone2": "연락처2",
    "seller": "특이사항",  # 셀러(비고) 값 — 매출과 동일
    "qty": "수량",
    "amount": "상품매입",
    "delivery": "매입배송비",
    "ferry": "매입도선료",  # 원본 '도선료.1' → load_orders 에서 rename
}

TITLE_SALE = "품목별판매현황"  # 셀러 시트 제목 (참고 정산서 row0과 동일)
# 2026-07-23 hoyeon.han: 매입 셀러 시트 제목 (경리나라 '품목별 구매 현황' 표기)
TITLE_PURCHASE = "품목별구매현황"
DETAIL_SHEET_NAME = "상세내역"  # 예약 시트명

# 2026-07-24 hoyeon.han: 상세내역 생성 모드 3택 (발주내역 기반 정산서생성에서 사용).
# 기본값은 SINGLE(현재 동작 = 전체 하나로) 이라 기존 셀러 화면과 하위호환.
DETAIL_MODE_NONE = "none"               # 생성 안 함
DETAIL_MODE_SINGLE = "single"           # 전체 하나로 (선택 셀러 전체를 상세 1장)
DETAIL_MODE_PER_SELLER = "per_seller"   # 셀러별 상세 시트 분리

ITEM_FERRY = "도선료"  # 도선료 라인 품목명
ITEM_PARCEL = "택배비"  # 판매배송비 라인 품목명

# 상세내역 시트 컬럼 순서 (논리명) — 출력 헤더는 column_map으로 원본명 표기
DETAIL_LOGICAL_COLS = [
    "date", "vendor", "category", "consignee", "address", "product",
    "phone1", "phone2", "seller", "qty", "amount", "delivery", "ferry",
]

_NUMERIC_LOGICAL_COLS = ["qty", "amount", "delivery", "ferry"]
_TEXT_LOGICAL_COLS = [
    "vendor", "category", "consignee", "address", "product",
    "phone1", "phone2", "seller",
]


# ===== 입력 파싱 =====


def load_orders(
    source: Any,
    sheet_name: str,
    column_map: dict[str, str] = COLUMN_MAP_SALE,
) -> pd.DataFrame:
    """발주내역 시트를 읽어 정산서 생성에 필요한 컬럼만 정규화해 반환.

    - 날짜: pd.Timestamp(0시 정규화). 파싱 실패는 NaT (필터에서 자연 제외)
    - 숫자: 콤마 등 제거 후 숫자 변환, NaN → 0
    - 문자: NaN → '' 후 strip
    - **어떤 행도 버리지 않는다** (수량 0 행 포함 — 필터/집계에서만 걸러짐)

    Args:
        source: 파일 경로(str) 또는 파일 객체
        sheet_name: 읽을 시트명 (예: '(누적)2026년 발주내역')
        column_map: 논리명 → 원본 컬럼명 매핑

    Raises:
        ValueError: 필수 컬럼 누락 시
    """
    df = read_order_sheet(source, sheet_name)

    # 2026-07-23 hoyeon.han: 발주내역은 '도선료' 헤더가 매출·매입 2개 → pandas 가
    # 두 번째를 '도선료.1'로 자동 접미한다. 매입 맵(COLUMN_MAP_PURCHASE)이 표준명
    # '매입도선료'로 참조하도록 정규화 (매출 경로는 첫 '도선료'만 쓰므로 무영향).
    if "도선료.1" in df.columns:
        df = df.rename(columns={"도선료.1": "매입도선료"})

    missing = [src for src in column_map.values() if src not in df.columns]
    if missing:
        raise ValueError(
            f"발주내역에 필요한 컬럼이 없습니다: {missing}. "
            f"시트('{sheet_name}')와 파일 양식을 확인해주세요."
        )

    out = df[list(column_map.values())].copy()

    date_col = column_map["date"]
    out[date_col] = pd.to_datetime(out[date_col], errors="coerce").dt.normalize()

    for key in _NUMERIC_LOGICAL_COLS:
        col = column_map[key]
        out[col] = to_num(out[col]).fillna(0)

    for key in _TEXT_LOGICAL_COLS:
        col = column_map[key]
        out[col] = out[col].fillna("").astype(str).str.strip()
        # astype(str) 과정에서 남는 'nan'/'None' 문자열 정리
        out[col] = out[col].replace({"nan": "", "None": ""})

    return out.reset_index(drop=True)


# ===== 선택지 추출 =====


def extract_order_dates(
    df: pd.DataFrame, column_map: dict[str, str] = COLUMN_MAP_SALE
) -> list[date]:
    """유효한 출고일 목록 (중복 제거, 오름차순). 개별 날짜 선택 모드용."""
    s = df[column_map["date"]].dropna()
    return sorted({ts.date() for ts in s})


def extract_date_range(
    df: pd.DataFrame, column_map: dict[str, str] = COLUMN_MAP_SALE
) -> tuple[date, date] | None:
    """유효 출고일의 (최소, 최대). 유효 날짜가 없으면 None."""
    dates = extract_order_dates(df, column_map)
    if not dates:
        return None
    return dates[0], dates[-1]


def extract_vendors(
    df: pd.DataFrame, column_map: dict[str, str] = COLUMN_MAP_SALE
) -> list[str]:
    """업체명 unique (빈 값 제외, 등장 순서 유지 — settlement.extract_vendors 패턴)."""
    values = df[column_map["vendor"]].fillna("").astype(str).str.strip()
    return [v for v in values.unique().tolist() if v]


def extract_sellers(
    df: pd.DataFrame, column_map: dict[str, str] = COLUMN_MAP_SALE
) -> list[str]:
    """셀러(특이사항) unique (빈 값 제외, 가나다순 — 선택 목록이 길어 정렬 표시)."""
    values = df[column_map["seller"]].fillna("").astype(str).str.strip()
    return sorted({v for v in values.unique().tolist() if v})


def extract_items(
    df: pd.DataFrame, column_map: dict[str, str] = COLUMN_MAP_SALE
) -> list[str]:
    """제품명 unique (빈 값 제외, 가나다순). 품목명 정리 편집기 입력용."""
    values = df[column_map["product"]].fillna("").astype(str).str.strip()
    return sorted({v for v in values.unique().tolist() if v})


# ===== 필터 =====


def filter_by_dates(
    df: pd.DataFrame,
    start_date: date | None = None,
    end_date: date | None = None,
    dates: Iterable[date] | None = None,
    column_map: dict[str, str] = COLUMN_MAP_SALE,
) -> pd.DataFrame:
    """일자 필터 (거래처 무관 — 거래처 선택지 추출용).

    일자 모드 (둘 중 하나 필수):
      - 기간 모드: start_date ~ end_date (inclusive)
      - 개별 날짜 모드: dates (해당 날짜들만)
    """
    date_col = column_map["date"]
    mask = df[date_col].notna()

    if dates is not None:
        wanted = {pd.Timestamp(d).normalize() for d in dates}
        if not wanted:
            raise ValueError("개별 날짜 모드에서는 날짜를 1개 이상 지정해야 합니다.")
        mask &= df[date_col].isin(list(wanted))
    elif start_date is not None and end_date is not None:
        mask &= (df[date_col] >= pd.Timestamp(start_date)) & (
            df[date_col] <= pd.Timestamp(end_date)
        )
    else:
        raise ValueError(
            "기간(start_date/end_date) 또는 개별 날짜(dates) 중 하나는 지정해야 합니다."
        )

    return df[mask].copy().reset_index(drop=True)


def filter_orders(
    df: pd.DataFrame,
    vendor: str,
    start_date: date | None = None,
    end_date: date | None = None,
    dates: Iterable[date] | None = None,
    column_map: dict[str, str] = COLUMN_MAP_SALE,
) -> pd.DataFrame:
    """일자 + 거래처 필터.

    계산서/수량/금액 조건은 걸지 않는다 (반품·수량 0 행 포함 —
    참고 정산서와 대조 검증된 규칙).
    """
    dated = filter_by_dates(df, start_date, end_date, dates, column_map)
    vendor_norm = dated[column_map["vendor"]].fillna("").astype(str).str.strip()
    return dated[vendor_norm == str(vendor).strip()].copy().reset_index(drop=True)


def filter_by_seller(
    df: pd.DataFrame,
    seller: str,
    column_map: dict[str, str] = COLUMN_MAP_SALE,
) -> pd.DataFrame:
    """셀러(특이사항) == seller 행만 반환."""
    seller_norm = df[column_map["seller"]].fillna("").astype(str).str.strip()
    return df[seller_norm == str(seller).strip()].copy().reset_index(drop=True)


def filter_by_sellers(
    df: pd.DataFrame,
    sellers: Iterable[str],
    column_map: dict[str, str] = COLUMN_MAP_SALE,
) -> pd.DataFrame:
    """셀러(특이사항)가 sellers 중 하나인 행만 반환 (선택 순서와 무관하게 원본 순서 유지)."""
    wanted = {str(s).strip() for s in sellers}
    seller_norm = df[column_map["seller"]].fillna("").astype(str).str.strip()
    return df[seller_norm.isin(wanted)].copy().reset_index(drop=True)


# ===== 셀러 시트 집계 =====


def aggregate_item_sheet(
    rows: pd.DataFrame,
    item_mapping: dict[str, str] | None = None,
    column_map: dict[str, str] = COLUMN_MAP_SALE,
) -> pd.DataFrame:
    """셀러 시트('품목별판매현황') 데이터 행 집계.

    참고 정산서와 대조 검증된 규칙:
      - 제품 라인: (출고일, 품목, 단가, 반품여부) 단위로 수량·금액 합산.
        단가 = 상품매출/수량 (행 단위). 수량 0 & 금액≠0 행은 단가 0 그룹으로
        포함 (금액이 합계에 반영됨 — 누락 금지)
      - 도선료 라인: 도선료>0 행을 (출고일, 도선료값) 그룹 →
        품목='도선료', 수량=행수, 단가=도선료값
      - 택배비 라인: 판매배송비>0 행 동일 규칙, 품목='택배비'
      - 정렬: 출고일 → 품목명 → 양수 먼저 반품(음수) 뒤

    Args:
        rows: filter_orders + filter_by_seller 를 거친 행들
        item_mapping: 품목명 정리 매핑 {원본 제품명: 표기명} (제품 라인에만 적용,
            도선료/택배비/상세내역에는 미적용)

    Returns:
        DataFrame[작성일자, 품목, 수량, 단가, 합계, 비고] — 데이터 행만
        (하단 합계행은 write_seller_settlement_xlsx 가 추가)
    """
    if rows.empty:
        return pd.DataFrame(columns=list(OUTPUT_COLS))

    c = column_map
    mapped = apply_mapping(rows, c["product"], item_mapping or {})

    work = pd.DataFrame(
        {
            "_날짜": mapped[c["date"]],
            "_품목": mapped[c["product"]].fillna("").astype(str).str.strip(),
            "_수량": pd.to_numeric(mapped[c["qty"]], errors="coerce").fillna(0.0),
            "_금액": pd.to_numeric(mapped[c["amount"]], errors="coerce").fillna(0.0),
            "_배송비": pd.to_numeric(mapped[c["delivery"]], errors="coerce").fillna(0.0),
            "_도선료": pd.to_numeric(mapped[c["ferry"]], errors="coerce").fillna(0.0),
            "_셀러": mapped[c["seller"]].fillna("").astype(str).str.strip(),
        }
    )

    qty = work["_수량"].to_numpy(dtype=float)
    amt = work["_금액"].to_numpy(dtype=float)
    # 행 단위 단가 = 금액/수량. 수량 0 행은 단가 0 그룹으로 묶되 금액은 보존.
    # (분모 0 회피를 위해 수량 0 자리에 1을 넣고 where로 결과만 0 처리)
    unit = np.where(qty != 0, amt / np.where(qty != 0, qty, 1.0), 0.0)
    work["_단가"] = np.round(unit, 4)  # float 노이즈로 그룹이 갈라지는 것 방지
    work["_반품"] = qty < 0

    group_cols = ["_날짜", "_품목", "_단가", "_반품", "_셀러"]
    prod = (
        work.groupby(group_cols, sort=False, dropna=False)
        .agg(_수량합=("_수량", "sum"), _합계=("_금액", "sum"))
        .reset_index()
    )

    # 도선료/택배비 라인 (해당 금액 > 0 인 행만 라인화 — 검증된 규칙)
    def _fee_lines(fee_col: str, item_name: str) -> pd.DataFrame:
        fee_rows = work[work[fee_col] > 0]
        if fee_rows.empty:
            return pd.DataFrame(columns=prod.columns)
        g = (
            fee_rows.groupby(["_날짜", fee_col, "_셀러"], sort=False)
            .size()
            .reset_index(name="_수량합")
            .rename(columns={fee_col: "_단가"})
        )
        g["_품목"] = item_name
        g["_반품"] = False
        g["_합계"] = g["_수량합"] * g["_단가"]
        return g[prod.columns.tolist()]

    # 2026-07-06 hoyeon.han: 도선료/택배비가 모두 없는 셀러는 _fee_lines가
    # 빈(all-NA) 프레임을 반환한다. 이를 concat에 넣으면 pandas가 _날짜 컬럼의
    # dtype을 datetime64 → object 로 강등해 이후 .dt.strftime 이 깨지므로
    # (pandas 2.2 FutureWarning → 3.0 AttributeError), 비어있지 않은 것만 합친다.
    # (rows.empty 는 함수 초입에서 조기 반환되고 groupby(dropna=False)라 prod는
    #  항상 1행 이상이므로 리스트가 완전히 비는 경우는 없다.)
    _parts = [
        prod,
        _fee_lines("_도선료", ITEM_FERRY),
        _fee_lines("_배송비", ITEM_PARCEL),
    ]
    merged = pd.concat(
        [p for p in _parts if not p.empty], ignore_index=True
    )

    merged = merged.sort_values(
        ["_날짜", "_품목", "_반품", "_단가"], kind="stable"
    ).reset_index(drop=True)

    out = pd.DataFrame(
        {
            "작성일자": merged["_날짜"].dt.strftime("%Y-%m-%d"),
            "품목": merged["_품목"],
            "수량": merged["_수량합"].round().astype(int),
            "단가": merged["_단가"],
            "합계": merged["_합계"].round().astype(int),
            "비고": merged["_셀러"],
        }
    )
    return out[list(OUTPUT_COLS)]


# ===== 상세내역 =====


def build_detail_df(
    df_filtered: pd.DataFrame,
    sellers: Iterable[str],
    item_mapping: dict[str, str] | None = None,
    column_map: dict[str, str] = COLUMN_MAP_SALE,
) -> pd.DataFrame:
    """상세내역 시트용 DataFrame (선택된 모든 셀러의 행).

    컬럼 순서는 참고 정산서 상세내역과 동일 (DETAIL_LOGICAL_COLS 순),
    헤더는 원본 컬럼명을 그대로 사용한다.

    2026-07-06 hoyeon.han: 제품(품목) 컬럼에도 품목명 정리(item_mapping)를 적용한다.
    (셀러 시트 집계와 동일한 명칭이 상세내역에도 표기되도록 — 사용자 요청)
    """
    rows = filter_by_sellers(df_filtered, sellers, column_map)
    if item_mapping:
        rows = apply_mapping(rows, column_map["product"], item_mapping)
    cols = [column_map[k] for k in DETAIL_LOGICAL_COLS]
    detail = rows[cols].copy()
    detail[column_map["date"]] = detail[column_map["date"]].dt.strftime("%Y-%m-%d")
    return detail.reset_index(drop=True)


# ===== 시트 계획 =====


def resolve_sheet_map(
    sellers: Iterable[str],
    seller_to_sheet: dict[str, str] | None = None,
) -> dict[str, list[str]]:
    """셀러 → 배치할 시트 매핑을 {시트명: [셀러...]} 로 뒤집는다.

    기존 정산서생성(pages/settlement.py)의 시트 배치 방식과 동일한 철학:
      - seller_to_sheet 에 없거나 값이 빈 문자열이면 셀러명 자신을 시트명으로.
      - 같은 시트명으로 여러 셀러를 묶으면 한 시트로 통합된다.
      - 시트명은 sanitize_sheet_name 규칙(31자, 금지문자) 적용.
    시트 순서는 sellers 등장 순서 기준 첫 등장 순.

    Raises:
        ValueError: 시트명이 예약어('상세내역')일 때
    """
    seller_to_sheet = seller_to_sheet or {}
    sheet_map: dict[str, list[str]] = {}
    for seller in sellers:
        s = str(seller).strip()
        raw = str(seller_to_sheet.get(seller, "")).strip() or s
        name = sanitize_sheet_name(raw)
        if name == DETAIL_SHEET_NAME:
            raise ValueError(
                f"시트명 '{DETAIL_SHEET_NAME}'은(는) 예약된 이름입니다. 다른 시트명을 입력해주세요."
            )
        sheet_map.setdefault(name, []).append(seller)
    return sheet_map


def _build_per_seller_detail(
    df_filtered: pd.DataFrame,
    seller_list: list[str],
    reserved_sheet_names: Iterable[str],
    item_mapping: dict[str, str] | None = None,
    column_map: dict[str, str] = COLUMN_MAP_SALE,
) -> dict[str, pd.DataFrame]:
    """셀러별 상세내역 시트 매핑 생성 (per_seller 모드) — 2026-07-24 hoyeon.han.

    각 셀러의 상세내역을 '상세내역_{셀러}' 시트로 분리한다. 시트배치로 여러
    셀러를 한 요약시트로 묶어도 상세는 셀러 단위로 나뉜다.

    - 시트명: sanitize_sheet_name('상세내역_{셀러}') (31자·금지문자 처리)
    - 중복(31자 절단 충돌 포함)은 '_2','_3' 접미로 회피
    - reserved_sheet_names(셀러 요약 시트명)와도 겹치지 않도록 seed
    - 데이터 없는 셀러는 건너뜀 (build_sheet_plan 의 빈 셀러 스킵과 동일)
    """
    used: set[str] = set(reserved_sheet_names)
    out: dict[str, pd.DataFrame] = {}
    for seller in seller_list:
        ddf = build_detail_df(df_filtered, [seller], item_mapping, column_map)
        if ddf.empty:
            continue
        base = sanitize_sheet_name(f"{DETAIL_SHEET_NAME}_{seller}")
        name = base
        n = 1
        while name in used:
            n += 1
            suffix = f"_{n}"
            name = base[: SHEET_MAX_LEN - len(suffix)] + suffix
        used.add(name)
        out[name] = ddf
    return out


def build_sheet_plan(
    df_filtered: pd.DataFrame,
    sellers: Iterable[str],
    seller_to_sheet: dict[str, str] | None = None,
    item_mapping: dict[str, str] | None = None,
    column_map: dict[str, str] = COLUMN_MAP_SALE,
    detail_mode: str = DETAIL_MODE_SINGLE,  # 2026-07-24 hoyeon.han: 상세내역 3택
) -> dict[str, Any]:
    """정산서 구성 계획 생성 (셀러 여러 개 → 시트 배치 매핑 지원).

    Args:
        sellers: 선택된 셀러(비고) 목록 (1개 이상)
        seller_to_sheet: {셀러: 배치할 시트명}. 빈 값이면 셀러명을 시트명으로 사용.
            같은 시트명으로 여러 셀러를 묶으면 통합 배치.
        item_mapping: 품목명 정리 매핑
        detail_mode: 상세내역 생성 모드 (DETAIL_MODE_NONE/SINGLE/PER_SELLER).
            기본 SINGLE = 현재 동작(전체 하나로). — 2026-07-24 hoyeon.han

    Returns:
        {
          'seller_sheets': {시트명: aggregate_item_sheet 결과 DataFrame},  # 배치 순서 유지
          'detail_mode':   전달된 detail_mode 값,
          'detail_df':     single 일 때 선택 셀러 전체 상세 DataFrame (그 외 None),
          'detail_sheets': per_seller 일 때 {'상세내역_{셀러}': DataFrame} (그 외 None),
        }

    Raises:
        ValueError: 셀러 미선택, 시트명 예약어 충돌, 데이터 없음, 알 수 없는 detail_mode
    """
    seller_list = [str(s).strip() for s in sellers if str(s).strip()]
    if not seller_list:
        raise ValueError("셀러(비고)를 1개 이상 선택해야 합니다.")

    sheet_map = resolve_sheet_map(seller_list, seller_to_sheet)

    seller_sheets: dict[str, pd.DataFrame] = {}
    for sheet_name, sheet_sellers in sheet_map.items():
        rows = filter_by_sellers(df_filtered, sheet_sellers, column_map)
        if rows.empty:
            continue  # 선택 조건에 데이터 없는 셀러는 빈 시트 대신 건너뜀
        seller_sheets[sheet_name] = aggregate_item_sheet(
            rows, item_mapping, column_map
        )

    if not seller_sheets:
        raise ValueError("선택한 조건에 해당하는 셀러 데이터가 없습니다.")

    # 2026-07-24 hoyeon.han: 상세내역 3택(detail_mode) 지원.
    # 기본 SINGLE 이면 기존과 동일하게 detail_df 를 담아 하위호환을 유지한다.
    # (기존 반환)
    # return {
    #     "seller_sheets": seller_sheets,
    #     "detail_df": build_detail_df(
    #         df_filtered, seller_list, item_mapping, column_map
    #     ),
    # }
    plan: dict[str, Any] = {
        "seller_sheets": seller_sheets,
        "detail_mode": detail_mode,
        "detail_df": None,
        "detail_sheets": None,
    }
    if detail_mode == DETAIL_MODE_SINGLE:
        plan["detail_df"] = build_detail_df(
            df_filtered, seller_list, item_mapping, column_map
        )
    elif detail_mode == DETAIL_MODE_PER_SELLER:
        plan["detail_sheets"] = _build_per_seller_detail(
            df_filtered, seller_list, list(seller_sheets.keys()),
            item_mapping, column_map,
        )
    elif detail_mode != DETAIL_MODE_NONE:
        raise ValueError(f"알 수 없는 상세내역 모드입니다: {detail_mode}")
    return plan


# ===== 헤더 텍스트 / 파일명 =====


def _normalize_dates(dates: Iterable[date]) -> list[date]:
    """날짜 목록 정규화 (중복 제거, 오름차순)."""
    return sorted({pd.Timestamp(d).date() for d in dates})


def build_period_line(
    vendor: str,
    start_date: date | None = None,
    end_date: date | None = None,
    dates: Iterable[date] | None = None,
) -> str:
    """셀러 시트 row 2 텍스트. 기간 모드는 settlement.build_period_line 재사용."""
    if dates is not None:
        ds = _normalize_dates(dates)
        if not ds:
            raise ValueError("날짜가 지정되지 않았습니다.")
        if len(ds) <= 3:
            date_part = ", ".join(d.isoformat() for d in ds)
        else:
            date_part = f"{ds[0].isoformat()} 외 {len(ds) - 1}일"
        return f"작성일자 : {date_part}    거래처 : {vendor}"
    if start_date is None or end_date is None:
        raise ValueError("기간(start_date/end_date) 또는 개별 날짜(dates)가 필요합니다.")
    return _build_period_line_range(start_date, end_date, vendor)


def build_output_filename(
    vendor: str,
    sheet_names: Iterable[str],
    doc_type: str = TYPE_SALE,
) -> str:
    """정산서 파일명 생성 (2026-07-06 hoyeon.han: 셀러 다중 선택 대응).

    - 시트가 1개(셀러 1개이거나 전부 한 시트로 통합)면:
        '정산서_{유형}_{거래처}_{시트명}.xlsx'  예: 정산서_매출_리버후드_드엘리사 4차.xlsx
    - 시트가 여러 개면 비고를 특정할 수 없으므로 생략:
        '정산서_{유형}_{거래처}.xlsx'           예: 정산서_매출_리버후드.xlsx
    (일자는 파일명에 포함하지 않는다)
    """
    names = [str(n) for n in sheet_names]
    doc_type_safe = _sanitize_filename_part(doc_type)
    vendor_safe = _sanitize_filename_part(vendor)
    if len(names) == 1:
        sheet_safe = _sanitize_filename_part(names[0])
        return f"정산서_{doc_type_safe}_{vendor_safe}_{sheet_safe}.xlsx"
    return f"정산서_{doc_type_safe}_{vendor_safe}.xlsx"


# ===== 출력 작성 =====


def _write_int_cell(ws, row: int, col: int, value: Any, font: Font | None = None):
    """정수 + '#,##0' 포맷 셀 작성 헬퍼."""
    cell = ws.cell(row=row, column=col, value=int(round(float(value))))
    cell.number_format = "#,##0"
    if font is not None:
        cell.font = font
    return cell


_HEADER_FILL = PatternFill(
    start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid"
)


def _write_seller_sheet(
    ws,
    seller_df: pd.DataFrame,
    title: str,
    period_line: str,
    seller_font_size: float,
    total_font_color: str,
    total_fill_color: str,
    total_bold: bool,
) -> None:
    """셀러 시트('품목별판매현황') 한 장을 작성한다.

    제목/기간/헤더/데이터/합계행 — 전체 셀 seller_font_size 적용,
    합계행은 굵게 + 합계(금액) 셀만 글자색/채우기색 강조.
    """
    base_font = Font(size=seller_font_size)
    header_font = Font(bold=True, size=seller_font_size)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    n_cols = len(OUTPUT_COLS)

    # 상단 3행
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = header_font
    cell.alignment = center
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    cell = ws.cell(row=2, column=1, value=period_line)
    cell.font = base_font
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)

    for c, col_name in enumerate(OUTPUT_COLS, start=1):
        cell = ws.cell(row=3, column=c, value=col_name)
        cell.font = header_font
        cell.fill = _HEADER_FILL
        cell.alignment = center

    # 컬럼 너비 (13pt 기준으로 settlement._apply_column_widths 대비 확대)
    seller_widths = {"작성일자": 14, "품목": 44, "수량": 12, "단가": 14, "합계": 17, "비고": 28}
    for c, name in enumerate(OUTPUT_COLS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = seller_widths.get(name, 16)

    # 데이터 행 (row 4+)
    for i, row in enumerate(seller_df.itertuples(index=False), start=4):
        for c, col_name in enumerate(OUTPUT_COLS, start=1):
            value = getattr(row, col_name, "")
            if col_name in ("수량", "합계"):
                _write_int_cell(ws, i, c, value, base_font)
            elif col_name == "단가":
                v = float(value) if value not in ("", None) else 0.0
                if v == int(v):
                    _write_int_cell(ws, i, c, v, base_font)
                else:
                    cell = ws.cell(row=i, column=c, value=v)
                    cell.number_format = "#,##0.####"
                    cell.font = base_font
            else:
                cell = ws.cell(row=i, column=c, value=str(value))
                cell.font = base_font

    # 합계행: ['', 'N건', 수량합, '', 금액합, ''] (참고 정산서 형식)
    last_row = 4 + len(seller_df)
    sum_qty = int(seller_df["수량"].sum()) if not seller_df.empty else 0
    sum_total = int(seller_df["합계"].sum()) if not seller_df.empty else 0

    ws.cell(row=last_row, column=1, value="")
    ws.cell(row=last_row, column=2, value=f"{len(seller_df)}건")
    qty_cell = _write_int_cell(ws, last_row, 3, sum_qty)
    ws.cell(row=last_row, column=4, value="")
    total_cell = _write_int_cell(ws, last_row, 5, sum_total)
    ws.cell(row=last_row, column=6, value="")
    qty_cell.alignment = right
    total_cell.alignment = right

    # 합계행 기본: 전체 셀 굵게 + 13pt (참고 정산서 합계행 스타일)
    row_bold_font = Font(bold=total_bold, size=seller_font_size)
    for c in range(1, n_cols + 1):
        ws.cell(row=last_row, column=c).font = row_bold_font

    # 2026-07-06 hoyeon.han: 색상 강조는 합계(금액) 셀만 적용한다.
    total_cell.font = Font(
        bold=total_bold,
        size=seller_font_size,
        color=_to_argb(total_font_color),
    )
    total_cell.fill = PatternFill(
        start_color=_to_argb(total_fill_color),
        end_color=_to_argb(total_fill_color),
        fill_type="solid",
    )


def _write_detail_sheet(
    ws,
    detail_df: pd.DataFrame,
    column_map: dict[str, str],
) -> None:
    """상세내역 시트를 작성한다 (원본 컬럼 + 열별 합계행 + 빈 2행 + 총합계행)."""
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    detail_cols = detail_df.columns.tolist()

    # 헤더 (row 1)
    for c, col_name in enumerate(detail_cols, start=1):
        cell = ws.cell(row=1, column=c, value=col_name)
        cell.font = bold
        cell.fill = _HEADER_FILL
        cell.alignment = center

    # 컬럼 너비 (논리명 기준 → 실제 컬럼명으로 변환)
    detail_width_by_logical = {
        "date": 12, "vendor": 16, "category": 12, "consignee": 12,
        "address": 44, "product": 38, "phone1": 15, "phone2": 15,
        "seller": 16, "qty": 8, "amount": 13, "delivery": 12, "ferry": 10,
    }
    detail_widths = {
        column_map[k]: w for k, w in detail_width_by_logical.items() if k in column_map
    }
    for c, name in enumerate(detail_cols, start=1):
        ws.column_dimensions[get_column_letter(c)].width = detail_widths.get(name, 14)

    qty_col = column_map["qty"]
    amount_col = column_map["amount"]
    delivery_col = column_map["delivery"]
    ferry_col = column_map["ferry"]

    # 데이터 행 (row 2+)
    for i, row in enumerate(detail_df.itertuples(index=False), start=2):
        for c, col_name in enumerate(detail_cols, start=1):
            value = row[c - 1]
            if col_name in (qty_col, amount_col):
                _write_int_cell(ws, i, c, value)
            elif col_name in (delivery_col, ferry_col):
                # 0은 빈칸 처리 (참고 정산서 상세내역과 동일)
                v = float(value) if value not in ("", None) else 0.0
                if v != 0:
                    _write_int_cell(ws, i, c, v)
                else:
                    ws.cell(row=i, column=c, value="")
            else:
                ws.cell(row=i, column=c, value=str(value))

    # 열별 합계행 → 빈 2행 → 총합계행 (참고 정산서 상세내역과 동일 구조)
    n_detail = len(detail_df)
    sums = {
        col: int(pd.to_numeric(detail_df[col], errors="coerce").fillna(0).sum())
        for col in (amount_col, delivery_col, ferry_col)
    }
    col_index = {name: idx + 1 for idx, name in enumerate(detail_cols)}

    total_row = 2 + n_detail
    for col, s in sums.items():
        _write_int_cell(ws, total_row, col_index[col], s, bold)

    grand_row = total_row + 3  # 빈 2행 뒤
    _write_int_cell(ws, grand_row, col_index[amount_col], sum(sums.values()), bold)

    ws.freeze_panes = "A2"


def write_seller_settlement_xlsx(
    out_path: str,
    title: str,
    period_line: str,
    plan: dict[str, Any],
    seller_font_size: float = 13.0,
    total_font_color: str = "#FF0000",
    total_fill_color: str = "#FFFF00",
    total_bold: bool = True,
    column_map: dict[str, str] = COLUMN_MAP_SALE,
) -> str:
    """셀러 정산서 .xlsx 작성 (2026-07-06 hoyeon.han: 셀러 다중 시트 지원).

    시트 구성 (요구사항 — '전체'/'일반' 시트 없음):
      1) 셀러 시트들: plan['seller_sheets'] 순서대로 (셀러 1개면 1장,
         비고를 시트별로 배치/통합한 만큼 여러 장). 전체 셀 seller_font_size,
         합계행은 굵게 + 합계 셀만 색상 강조.
      2) 상세내역 시트 (2026-07-24 hoyeon.han: plan['detail_mode'] 로 0/1/N장):
         - none: 생성 안 함
         - single(기본): 선택된 모든 셀러의 원본 행 1장('상세내역')
         - per_seller: 셀러별 '상세내역_{셀러}' N장 (셀러 시트 뒤)
         각 상세시트 = 원본 행 + 열별 합계행 + 빈 2행 + 총합계행 (기본 폰트 크기)
    """
    seller_sheets: dict[str, pd.DataFrame] = plan["seller_sheets"]
    # 2026-07-24 hoyeon.han: 상세내역 3택 지원 — detail_mode 를 읽어 상세시트 0/1/N장.
    # 키 없는 수동 plan 은 SINGLE 로 폴백(하위호환).
    # detail_df: pd.DataFrame = plan["detail_df"]
    detail_mode = plan.get("detail_mode", DETAIL_MODE_SINGLE)
    if not seller_sheets:
        raise ValueError("생성할 셀러 시트가 없습니다.")
    for name in seller_sheets:
        if name == DETAIL_SHEET_NAME:  # build_sheet_plan에서 검증되지만 방어
            raise ValueError(f"시트명 '{DETAIL_SHEET_NAME}'은(는) 예약된 이름입니다.")

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)

    wb = Workbook()

    # ---------- 셀러 시트들 ----------
    first = True
    for sheet_name, seller_df in seller_sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        _write_seller_sheet(
            ws, seller_df, title, period_line, seller_font_size,
            total_font_color, total_fill_color, total_bold,
        )

    # ---------- 상세내역 시트 (2026-07-24 hoyeon.han: 모드별 0/1/N장) ----------
    # (기존: 항상 1장)
    # ws_detail = wb.create_sheet(title=DETAIL_SHEET_NAME)
    # _write_detail_sheet(ws_detail, detail_df, column_map)
    if detail_mode == DETAIL_MODE_NONE:
        pass  # 상세내역 생성 안 함
    elif detail_mode == DETAIL_MODE_PER_SELLER:
        for sheet_name, ddf in (plan.get("detail_sheets") or {}).items():
            _write_detail_sheet(
                wb.create_sheet(title=sheet_name), ddf, column_map
            )
    else:  # SINGLE (또는 키 없는 수동 plan) → 현재 동작
        _write_detail_sheet(
            wb.create_sheet(title=DETAIL_SHEET_NAME), plan["detail_df"], column_map
        )

    wb.save(out_path)
    return out_path
