# order_file_store.py
# 2026-06-03 hoyeon.han: 발주내역 파일 서버 저장/재사용 기능 (신규)
#
# 발주내역 Excel(.xlsm) 파일을 서버에 1개 버전으로 영속 저장하고,
# 여러 페이지에서 재사용/다운로드할 수 있도록 관리한다.
#
# - 순수 모듈: Streamlit에 의존하지 않으며 bytes/str/dict만 입출력한다.
# - 저장 위치: order_data/current.xlsm (고정 파일명, 덮어쓰기로 1버전 유지)
# - 메타데이터: order_data/current.meta.json (원본파일명/업로드시각/크기/시트목록)
# - uploads/ 가 아닌 별도 디렉토리를 쓰는 이유: 시스템관리 페이지의 데이터 정리/
#   전체 초기화 기능이 uploads/ 안의 모든 파일을 삭제하기 때문.

import os
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Tuple, Optional, List

import openpyxl


class OrderFileStore:
    """발주내역 파일 서버 저장소 관리 클래스 (단일 버전 유지)

    기능:
    - save(): 업로드 파일 저장 (기존 파일 덮어쓰기 = 1버전 유지) + 메타 기록
    - exists()/get_path()/get_metadata()/get_file_bytes(): 조회
    - delete(): 파일 + 메타 삭제
    - get_stats(): 관리 UI용 요약 정보
    """

    def __init__(self, store_dir: str = "order_data"):
        """초기화

        Args:
            store_dir: 발주내역 파일 저장 디렉토리 경로
        """
        self.store_dir = Path(store_dir)
        self.file_path = self.store_dir / "current.xlsm"  # 고정 파일명 (1버전 유지)
        self.meta_path = self.store_dir / "current.meta.json"
        self.logger = logging.getLogger(__name__)

        # 저장 디렉토리 생성 (CustomerMasterDB 패턴)
        self.store_dir.mkdir(parents=True, exist_ok=True)

    # ==========================================================================
    # 저장
    # ==========================================================================

    def save(self, file_bytes: bytes, original_name: str) -> Tuple[bool, str]:
        """업로드된 발주내역 파일을 서버에 저장한다.

        기존 파일이 있으면 덮어써서 항상 1개 버전만 유지한다.
        저장 후 시트 목록을 추출하여 메타데이터(JSON)에 함께 기록한다.

        Args:
            file_bytes: 업로드 파일의 바이트 (Streamlit uploaded_file.getvalue())
            original_name: 원본 파일명 (uploaded_file.name)

        Returns:
            (성공 여부, 메시지)
        """
        try:
            # 파일 본체 저장 (덮어쓰기 = 기존 버전 자동 대체)
            with open(self.file_path, "wb") as f:
                f.write(file_bytes)

            # 시트 목록 추출 (실패해도 저장 자체는 성공 처리)
            sheet_names = self._extract_sheet_names()

            meta = {
                "original_name": original_name,
                "uploaded_at": datetime.now().isoformat(timespec="seconds"),
                "size_bytes": len(file_bytes),
                "sheet_names": sheet_names,
                "recommended_sheet": self._pick_recommended_sheet(sheet_names),
            }

            with open(self.meta_path, "w", encoding="utf-8") as f:
                json.dump(meta, f, ensure_ascii=False, indent=2)

            self.logger.info(
                f"발주내역 파일 저장 완료: {original_name} ({len(file_bytes):,} bytes)"
            )
            return True, f"발주내역 파일이 저장되었습니다: {original_name}"

        except Exception as e:
            self.logger.error(f"발주내역 파일 저장 실패: {e}", exc_info=True)
            return False, f"발주내역 파일 저장에 실패했습니다: {str(e)}"

    # ==========================================================================
    # 조회
    # ==========================================================================

    def exists(self) -> bool:
        """저장된 발주내역 파일이 있는지 여부."""
        return self.file_path.exists()

    def get_path(self) -> Optional[str]:
        """저장된 파일의 경로 문자열. 없으면 None.

        처리 함수(get_sales_by_period 등)에 그대로 넘길 수 있다.
        """
        return str(self.file_path) if self.exists() else None

    def get_metadata(self) -> Optional[dict]:
        """저장된 파일의 메타데이터(dict). 없거나 손상 시 None."""
        if not self.meta_path.exists():
            return None
        try:
            with open(self.meta_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError) as e:
            self.logger.warning(f"메타데이터 읽기 실패: {e}")
            return None

    def get_file_bytes(self) -> Optional[bytes]:
        """저장된 파일의 바이트. 다운로드 버튼 data=용. 없으면 None."""
        if not self.exists():
            return None
        with open(self.file_path, "rb") as f:
            return f.read()

    # ==========================================================================
    # 삭제
    # ==========================================================================

    def delete(self) -> Tuple[bool, str]:
        """저장된 파일과 메타데이터를 삭제한다 (멱등)."""
        try:
            removed = False
            for p in (self.file_path, self.meta_path):
                if p.exists():
                    p.unlink()
                    removed = True
            if removed:
                self.logger.info("발주내역 파일 삭제 완료")
                return True, "저장된 발주내역 파일을 삭제했습니다."
            return True, "삭제할 발주내역 파일이 없습니다."
        except Exception as e:
            self.logger.error(f"발주내역 파일 삭제 실패: {e}", exc_info=True)
            return False, f"삭제에 실패했습니다: {str(e)}"

    # ==========================================================================
    # 유틸
    # ==========================================================================

    def get_stats(self) -> dict:
        """관리 UI 표시용 요약 정보.

        Returns:
            {"exists": bool, "original_name": str, "uploaded_at": str,
             "size_kb": float, "sheet_count": int, "sheet_names": list,
             "recommended_sheet": str|None}
        """
        if not self.exists():
            return {"exists": False}
        meta = self.get_metadata() or {}
        size_bytes = meta.get("size_bytes")
        if size_bytes is None:
            size_bytes = self.file_path.stat().st_size
        return {
            "exists": True,
            "original_name": meta.get("original_name", self.file_path.name),
            "uploaded_at": meta.get("uploaded_at", ""),
            "size_kb": size_bytes / 1024,
            "sheet_count": len(meta.get("sheet_names", [])),
            "sheet_names": meta.get("sheet_names", []),
            "recommended_sheet": meta.get("recommended_sheet"),
        }

    def _extract_sheet_names(self) -> List[str]:
        """저장된 파일에서 시트 목록을 추출한다.

        pages/1_📝_전표생성.py 의 시트 검증 로직과 동일하게 read_only 모드 사용.
        실패 시 빈 리스트 반환.
        """
        try:
            wb = openpyxl.load_workbook(
                self.file_path, read_only=True, keep_links=False
            )
            names = list(wb.sheetnames)
            wb.close()
            return names
        except Exception as e:
            self.logger.warning(f"시트 목록 추출 실패: {e}")
            return []

    @staticmethod
    def _pick_recommended_sheet(sheet_names: List[str]) -> Optional[str]:
        """'발주내역' + 현재 연도를 포함한 시트를 우선 추천한다.

        pages/5_📆_전표생성_기간.py 의 기본 시트 선택 로직을 차용.
        매칭 실패 시 첫 번째 시트, 시트가 없으면 None.
        """
        if not sheet_names:
            return None
        current_year = str(datetime.now().year)
        for sheet in sheet_names:
            if "발주내역" in sheet and current_year in sheet:
                return sheet
        return sheet_names[0]


# ==============================================================================
# 스모크 테스트 (단독 실행: python Src/order_file_store.py)
# ==============================================================================

if __name__ == "__main__":
    import io
    import tempfile

    logging.basicConfig(level=logging.INFO)

    with tempfile.TemporaryDirectory() as tmp:
        store = OrderFileStore(store_dir=tmp)
        assert store.exists() is False, "초기에는 파일이 없어야 함"
        assert store.get_path() is None
        assert store.get_metadata() is None
        assert store.get_stats() == {"exists": False}

        # 테스트용 워크북 생성 (openpyxl 은 확장자와 무관하게 내용으로 읽음)
        year = datetime.now().year
        wb = openpyxl.Workbook()
        wb.active.title = f"(누적){year}년 발주내역"
        wb.create_sheet("기타시트")
        bio = io.BytesIO()
        wb.save(bio)

        ok, msg = store.save(bio.getvalue(), "테스트_발주내역.xlsm")
        print("save:", ok, msg)
        assert ok is True
        assert store.exists() is True
        assert store.get_path() == str(store.file_path)

        meta = store.get_metadata()
        print("meta:", meta)
        assert meta["original_name"] == "테스트_발주내역.xlsm"
        assert meta["recommended_sheet"] == f"(누적){year}년 발주내역"
        assert "기타시트" in meta["sheet_names"]

        stats = store.get_stats()
        print("stats:", stats)
        assert stats["exists"] is True
        assert stats["sheet_count"] == 2

        assert store.get_file_bytes() == bio.getvalue()

        ok, msg = store.delete()
        print("delete:", ok, msg)
        assert ok is True
        assert store.exists() is False

        print("✅ 모든 스모크 테스트 통과")
