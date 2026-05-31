# 2026-05-27 hoyeon.han 정산서 생성 비즈니스 로직 신규 작성
"""정산서(매출/매입) 생성 모듈.

회계 시스템에서 다운로드한 "기간 내 전체 거래처" 단일 .xls를 입력으로 받아,
거래처/기간을 지정하고 품목/비고를 정제한 뒤, 비고별로 시트가 분리된
.xlsx 정산서를 생성한다.

본 모듈의 모든 함수는 순수 함수(Streamlit/세션상태 의존 없음)로 작성한다.
"""

from __future__ import annotations

import os
import re
from datetime import date, datetime
from typing import Any

import pandas as pd
import xlrd  # .xls 입력 파싱 (xlrd 2.0+)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ===== 상수 =====

DELETE_COLS = ["회사명", "공급가", "부가세"]
OUTPUT_COLS = ["작성일자", "품목", "수량", "단가", "합계", "비고"]
SHEET_OVERALL = "전체"  # 첫 번째 시트 — 전체 데이터 + 특수 합계 포맷
SHEET_GENERAL = "일반"  # 두 번째 시트 — 비고 비어있는 행
TYPE_SALE = "매출"
TYPE_PURCHASE = "매입"

# row 1 col A 텍스트에서 기간 추출 — `_:_` / ` : ` / `___:___` 등 변형 허용
PERIOD_RE = re.compile(
    r"작성일자[\s_:]+(\d{4}-\d{2}-\d{2})[\s_~]+(\d{4}-\d{2}-\d{2})"
)

# 엑셀 시트명 금지문자 / 길이 제한
SHEET_INVALID_RE = re.compile(r"[:\\/?*\[\]]")
SHEET_MAX_LEN = 31

# 파일명 금지문자(OS 공통)
FILENAME_INVALID_RE = re.compile(r'[/\\:*?"<>|]')


# ===== 입력 파싱 =====


def _cell(sheet: xlrd.sheet.Sheet, row: int, col: int) -> Any:
    """xlrd 셀 안전 조회 (범위 밖이면 빈 문자열)."""
    if row >= sheet.nrows or col >= sheet.ncols:
        return ""
    return sheet.cell_value(row, col)


def _detect_doc_type(title: str) -> str | None:
    """row 0 제목 텍스트로 매출/매입 추정. 미상이면 None."""
    if not isinstance(title, str):
        return None
    if "판매" in title:
        return TYPE_SALE
    if "구매" in title:
        return TYPE_PURCHASE
    return None


def _parse_period(period_text: str) -> tuple[date, date]:
    """row 1 텍스트에서 시작/종료일 추출."""
    if not isinstance(period_text, str):
        raise ValueError(f"기간 정보를 인식할 수 없습니다: {period_text!r}")
    m = PERIOD_RE.search(period_text)
    if not m:
        raise ValueError(
            f"기간 정보를 인식할 수 없습니다(형식 불일치): {period_text!r}"
        )
    start = datetime.strptime(m.group(1), "%Y-%m-%d").date()
    end = datetime.strptime(m.group(2), "%Y-%m-%d").date()
    return start, end


def _normalize_date_column(series: pd.Series) -> pd.Series:
    """작성일자 컬럼을 YYYY-MM-DD 문자열로 정규화. 변환 실패 시 원본 유지."""
    parsed = pd.to_datetime(series, errors="coerce")
    out = parsed.dt.strftime("%Y-%m-%d")
    # 변환 실패한 셀은 원본 문자열로 복구
    return out.where(parsed.notna(), series.astype(str))


def parse_settlement_xls(file_path_or_buffer: Any) -> dict[str, Any]:
    """정산서 원본 .xls 파싱.

    Returns:
        {
          'title':       row 0 col A (예: '품목별판매현황'),
          'period_text': row 1 col A 원문,
          'start_date':  date,
          'end_date':    date,
          'doc_type':    '매출' | '매입' | None,
          'df':          DataFrame (row 2 헤더 기준, 회사명='합계'/빈 행 제외),
        }
    """
    # xlrd는 파일 경로 또는 file_contents=bytes를 지원
    if hasattr(file_path_or_buffer, "read"):
        contents = file_path_or_buffer.read()
        book = xlrd.open_workbook(file_contents=contents)
    else:
        book = xlrd.open_workbook(file_path_or_buffer)

    sheet = book.sheet_by_index(0)
    if sheet.nrows < 4:
        raise ValueError("정산서 원본 파일의 행 수가 부족합니다.")

    title = str(_cell(sheet, 0, 0)).strip()
    period_text = str(_cell(sheet, 1, 0)).strip()
    start_date, end_date = _parse_period(period_text)
    doc_type = _detect_doc_type(title)

    # row 2 = 헤더, row 3+ = 데이터
    headers = [str(_cell(sheet, 2, c)).strip() for c in range(sheet.ncols)]
    data_rows = []
    for r in range(3, sheet.nrows):
        row = [_cell(sheet, r, c) for c in range(sheet.ncols)]
        data_rows.append(row)

    df = pd.DataFrame(data_rows, columns=headers)

    # 회사명 == '합계' 또는 빈 행 제거
    if "회사명" in df.columns:
        df = df[df["회사명"].astype(str).str.strip() != "합계"].copy()
        df = df[df["회사명"].astype(str).str.strip() != ""].copy()

    # 작성일자 컬럼 정규화 (엑셀 시리얼 / datetime / 문자열 혼재 대응)
    if "작성일자" in df.columns:
        df["작성일자"] = _normalize_date_column(df["작성일자"])

    return {
        "title": title,
        "period_text": period_text,
        "start_date": start_date,
        "end_date": end_date,
        "doc_type": doc_type,
        "df": df.reset_index(drop=True),
    }


# ===== 필터링 / 변환 =====


def extract_vendors(df: pd.DataFrame) -> list[str]:
    """회사명 unique 목록을 등장 순서로 반환. '합계'/빈 값 제외."""
    if "회사명" not in df.columns:
        return []
    seen: list[str] = []
    seen_set: set[str] = set()
    for v in df["회사명"].astype(str).str.strip():
        if v in ("", "합계"):
            continue
        if v not in seen_set:
            seen.append(v)
            seen_set.add(v)
    return seen


def filter_by_vendor(df: pd.DataFrame, vendor: str) -> pd.DataFrame:
    """회사명 == vendor 필터링."""
    if "회사명" not in df.columns:
        return df.copy()
    return df[df["회사명"].astype(str).str.strip() == vendor].copy()


def filter_by_period(
    df: pd.DataFrame, start_date: date, end_date: date
) -> pd.DataFrame:
    """작성일자가 [start_date, end_date] 범위 내인 행만 반환."""
    if "작성일자" not in df.columns:
        return df.copy()
    parsed = pd.to_datetime(df["작성일자"], errors="coerce").dt.date
    mask = (parsed >= start_date) & (parsed <= end_date)
    # 날짜 파싱 실패한 행은 안전하게 제외
    return df[mask.fillna(False)].copy()


def drop_unused_columns(df: pd.DataFrame) -> pd.DataFrame:
    """DELETE_COLS 제거 후 OUTPUT_COLS 순서로 재정렬."""
    keep = [c for c in df.columns if c not in DELETE_COLS]
    out = df[keep].copy()
    # OUTPUT_COLS 중 누락된 컬럼이 있으면 빈 컬럼 추가
    for c in OUTPUT_COLS:
        if c not in out.columns:
            out[c] = ""
    return out[OUTPUT_COLS].copy()


def apply_mapping(
    df: pd.DataFrame, column: str, mapping: dict[str, str]
) -> pd.DataFrame:
    """column 값을 mapping에 따라 치환. mapping에 없으면 원본 유지."""
    if column not in df.columns or not mapping:
        return df.copy()
    out = df.copy()
    out[column] = out[column].astype(str).map(lambda v: mapping.get(v, v))
    return out


# ===== 숫자 정규화 =====


def _to_int_safe(series: pd.Series) -> pd.Series:
    """숫자 컬럼을 int로 안전 변환. 콤마/공백 제거 후 NaN은 0."""
    s = series.astype(str).str.strip().str.replace(",", "", regex=False)
    num = pd.to_numeric(s, errors="coerce").fillna(0)
    return num.astype(int)


# ===== 시트 분할 =====


def split_into_sheets(
    df: pd.DataFrame,
    remark_to_sheet: dict[str, str] | None = None,
) -> dict[str, pd.DataFrame]:
    """비고 기반 시트 분할.

    비고 컬럼의 데이터 값은 절대 변경하지 않고, `remark_to_sheet` 매핑을 통해
    각 행이 배치될 시트만 결정한다. (예: 비고 "쏘피 2차"를 "쏘피" 시트로 묶고
    싶을 때 사용. 행 자체의 비고 값은 "쏘피 2차"로 유지됨.)

    Args:
        df: drop_unused_columns 적용 후 DataFrame.
        remark_to_sheet: {원본 비고: 배치할 시트명}.
            - 키에 없는 비고는 자기 자신을 시트명으로 사용.
            - 값이 빈 문자열 또는 '일반'이면 '일반' 시트로 분류.
            - 비고가 빈 값인 행은 항상 '일반' 시트.

    Returns:
        키 순서: ['전체', '일반', *sorted(other_sheets, key=str.casefold)]
        각 값: 해당 시트에 들어갈 DataFrame (.copy()).
    """
    if "비고" not in df.columns:
        raise ValueError("비고 컬럼이 없어 시트 분할이 불가합니다.")

    if remark_to_sheet is None:
        remark_to_sheet = {}

    remark_norm = df["비고"].fillna("").astype(str).str.strip()

    def _resolve_target(remark: str) -> str:
        """원본 비고 → 배치 시트명."""
        if remark == "":
            return SHEET_GENERAL
        target = str(remark_to_sheet.get(remark, remark)).strip()
        if target == "" or target == SHEET_GENERAL:
            return SHEET_GENERAL
        return target

    sheet_target = remark_norm.map(_resolve_target)

    sheets: dict[str, pd.DataFrame] = {}
    sheets[SHEET_OVERALL] = df.copy().reset_index(drop=True)
    sheets[SHEET_GENERAL] = (
        df[sheet_target == SHEET_GENERAL].copy().reset_index(drop=True)
    )

    other_sheet_names = sorted(
        {
            s
            for s in sheet_target.unique()
            if s not in (SHEET_OVERALL, SHEET_GENERAL)
        },
        key=lambda s: s.casefold(),
    )
    for sheet_name in other_sheet_names:
        sheets[sheet_name] = (
            df[sheet_target == sheet_name].copy().reset_index(drop=True)
        )

    return sheets


# ===== 이름 sanitize =====


def sanitize_sheet_name(name: str) -> str:
    """엑셀 시트명 규칙 적용: 31자, 금지문자 '_' 치환, 빈문자열 '_'."""
    s = SHEET_INVALID_RE.sub("_", str(name))
    if len(s) > SHEET_MAX_LEN:
        s = s[:SHEET_MAX_LEN]
    return s or "_"


def _sanitize_filename_part(name: str) -> str:
    """파일명 금지문자 '_' 치환."""
    return FILENAME_INVALID_RE.sub("_", str(name)).strip() or "_"


def build_period_line(
    start_date: date, end_date: date, vendor: str
) -> str:
    """정산서 상단 row 1 텍스트 생성."""
    return (
        f"작성일자 : {start_date.isoformat()} ~ {end_date.isoformat()}"
        f"    거래처 : {vendor}"
    )


def build_output_filename(
    doc_type: str, vendor: str, start_date: date, end_date: date
) -> str:
    """'정산서_매출_<vendor>_<start>_<end>.xlsx' 형식."""
    vendor_safe = _sanitize_filename_part(vendor)
    doc_type_safe = _sanitize_filename_part(doc_type)
    return (
        f"정산서_{doc_type_safe}_{vendor_safe}"
        f"_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"
    )


# ===== 출력 작성 =====


# 2026-05-31 hoyeon.han: st.color_picker가 반환하는 '#RRGGBB'를 openpyxl이
# 기대하는 'AARRGGBB'(ARGB) 형식으로 정규화한다.
def _to_argb(color: str) -> str:
    """'#RRGGBB' / 'RRGGBB' / 'AARRGGBB' → 'AARRGGBB'. 알파 누락 시 'FF' 선행."""
    s = str(color).lstrip("#").strip().upper()
    if len(s) == 6:
        s = "FF" + s
    return s


def _write_common_header(
    ws,
    title: str,
    period_line: str,
    header_cols: list[str],
    bold_font: Font,
    header_fill: PatternFill,
    center: Alignment,
) -> None:
    """모든 시트 공통 상단 3행 작성."""
    n_cols = len(header_cols)
    # row 1: title (A1만), A1:F1 병합
    ws.cell(row=1, column=1, value=title).font = bold_font
    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=n_cols
    )
    ws.cell(row=1, column=1).alignment = center

    # row 2: 기간/거래처 (A2만), A2:F2 병합
    ws.cell(row=2, column=1, value=period_line)
    ws.merge_cells(
        start_row=2, start_column=1, end_row=2, end_column=n_cols
    )
    # 2026-05-31 hoyeon.han: 2행(작성일자/거래처) 좌측 정렬로 변경
    # ws.cell(row=2, column=1).alignment = center
    ws.cell(row=2, column=1).alignment = Alignment(
        horizontal="left", vertical="center"
    )

    # row 3: 헤더
    for c, col_name in enumerate(header_cols, start=1):
        cell = ws.cell(row=3, column=c, value=col_name)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center


def _apply_column_widths(ws, header_cols: list[str]) -> None:
    """가독성을 위한 기본 컬럼 너비."""
    widths = {
        "작성일자": 12,
        "품목": 38,
        "수량": 10,
        "단가": 12,
        "합계": 14,
        "비고": 24,
    }
    for c, name in enumerate(header_cols, start=1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(name, 14)


def write_settlement_xlsx(
    out_path: str,
    title: str,
    period_line: str,
    sheets: dict[str, pd.DataFrame],
    # 2026-05-31 hoyeon.han: 전체 시트 금액 합계 강조 옵션 (사용자 지정)
    highlight_total: bool = True,
    total_font_size: float = 11,
    total_font_color: str = "#FF0000",
    total_fill_color: str = "#FFFF00",
    total_bold: bool = True,
) -> str:
    """정산서 .xlsx 작성.

    각 시트 공통:
      A1 = title, A2 = period_line, A3:F3 = OUTPUT_COLS
      데이터 row 4+
    하단 합계 행:
      - '전체' 시트: ['', f'{N}건', f'{int(sum_qty):,}', '', f'{int(sum_total):,}', '']  (문자열, 콤마)
      - 그 외:      ['합계', '', sum_qty, '', sum_total, '']  (숫자)

    수량/합계 컬럼은 number_format='#,##0'.
    """
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)

    wb = Workbook()
    # Workbook은 기본 시트가 1개 있음 → 첫 시트를 첫 sheet name으로 사용 후 나머지 추가
    default_ws = wb.active
    bold_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid"
    )
    center = Alignment(horizontal="center", vertical="center")
    # 2026-05-31 hoyeon.han: 전체 시트 합계(수량/금액) 우측 정렬용
    right = Alignment(horizontal="right", vertical="center")
    # 2026-05-31 hoyeon.han: 금액 합계 강조 스타일 (옵션 켜짐일 때만 사용)
    total_font = Font(
        bold=total_bold,
        color=_to_argb(total_font_color),
        size=total_font_size,
    )
    total_fill = PatternFill(
        start_color=_to_argb(total_fill_color),
        end_color=_to_argb(total_fill_color),
        fill_type="solid",
    )

    used_names: set[str] = set()

    for idx, (raw_name, df_sheet) in enumerate(sheets.items()):
        sheet_name = sanitize_sheet_name(raw_name)
        # 시트명 중복 방지
        base = sheet_name
        suffix = 1
        while sheet_name in used_names:
            suffix += 1
            candidate = f"{base[: SHEET_MAX_LEN - 3]}_{suffix}"
            sheet_name = candidate[:SHEET_MAX_LEN]
        used_names.add(sheet_name)

        if idx == 0:
            ws = default_ws
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        _write_common_header(
            ws, title, period_line, OUTPUT_COLS,
            bold_font, header_fill, center,
        )
        _apply_column_widths(ws, OUTPUT_COLS)

        # 숫자 컬럼 정규화 (한 번만)
        qty = _to_int_safe(df_sheet.get("수량", pd.Series(dtype=object)))
        total = _to_int_safe(df_sheet.get("합계", pd.Series(dtype=object)))

        # 데이터 작성 (row 4+)
        for i, row in enumerate(df_sheet.itertuples(index=False), start=4):
            for c, col_name in enumerate(OUTPUT_COLS, start=1):
                value = getattr(row, col_name, "")
                if col_name == "수량":
                    cell = ws.cell(row=i, column=c, value=int(qty.iloc[i - 4]))
                    cell.number_format = "#,##0"
                elif col_name == "합계":
                    cell = ws.cell(
                        row=i, column=c, value=int(total.iloc[i - 4])
                    )
                    cell.number_format = "#,##0"
                elif col_name == "단가":
                    # 단가는 0/숫자가 섞일 수 있으므로 안전 변환
                    try:
                        v_num = int(float(value)) if value not in ("", None) else ""
                        cell = ws.cell(row=i, column=c, value=v_num)
                        if isinstance(v_num, int):
                            cell.number_format = "#,##0"
                    except (TypeError, ValueError):
                        ws.cell(row=i, column=c, value=str(value))
                else:
                    ws.cell(row=i, column=c, value=value)

        # 합계 행 작성
        last_row = 4 + len(df_sheet)
        sum_qty = int(qty.sum())
        sum_total = int(total.sum())

        if raw_name == SHEET_OVERALL:
            # 특수 포맷 — 문자열, 콤마
            ws.cell(row=last_row, column=1, value="")
            ws.cell(row=last_row, column=2, value=f"{len(df_sheet)}건")
            qty_overall_cell = ws.cell(
                row=last_row, column=3, value=f"{sum_qty:,}"
            )
            ws.cell(row=last_row, column=4, value="")
            total_overall_cell = ws.cell(
                row=last_row, column=5, value=f"{sum_total:,}"
            )
            ws.cell(row=last_row, column=6, value="")
            # 2026-05-31 hoyeon.han: 수량/금액 합계 우측 정렬
            qty_overall_cell.alignment = right
            total_overall_cell.alignment = right
        else:
            ws.cell(row=last_row, column=1, value="합계")
            ws.cell(row=last_row, column=2, value="")
            qty_cell = ws.cell(row=last_row, column=3, value=sum_qty)
            qty_cell.number_format = "#,##0"
            ws.cell(row=last_row, column=4, value="")
            total_cell = ws.cell(row=last_row, column=5, value=sum_total)
            total_cell.number_format = "#,##0"
            ws.cell(row=last_row, column=6, value="")

        # 합계 행 굵게
        for c in range(1, len(OUTPUT_COLS) + 1):
            ws.cell(row=last_row, column=c).font = bold_font

        # 2026-05-31 hoyeon.han: 전체 시트 금액 합계 강조 옵션
        # (사용자 지정 폰트 크기/글씨 색상/채우기 색상). 굵게 루프 이후에
        # 적용해 font가 덮어쓰이지 않게 한다. alignment는 font/fill과 독립이라
        # 우측 정렬은 그대로 유지된다.
        if raw_name == SHEET_OVERALL and highlight_total:
            total_cell = ws.cell(row=last_row, column=5)
            total_cell.font = total_font
            total_cell.fill = total_fill

    wb.save(out_path)
    return out_path
